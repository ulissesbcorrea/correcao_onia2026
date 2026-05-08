"""XLSX Parser for ONIA exam data.

Columns: Id | Avaliador | Status | Justificativa | Nome Completo | Acertos |
         Estado | Município | Escola | Polo | Situação | Q1..Q20

Yellow-filled rows = high fraud risk.
"""

import hashlib
import openpyxl
from openpyxl.utils import get_column_letter
from app.extensions import db
from app.models import Student, School, Answer, Justification, FraudFlag, ImportLog


# Column indices (0-based) matching the XLSX structure
COLUMN_MAP = {
    "xlsx_id": 0,       # Id
    "avaliador": 1,     # Avaliador
    "status": 2,        # Status
    "justificativa": 3, # Justificativa
    "nome": 4,          # Nome Completo
    "acertos": 5,       # Acertos
    "estado": 6,        # Estado
    "municipio": 7,     # Município
    "escola": 8,        # Escola
    "polo": 9,          # Polo
    "situacao": 10,     # Situação
    "q1": 11,           # Q1
}


def is_yellow_fill(cell):
    """Check if a cell has yellow fill (high fraud risk marker)."""
    fill = cell.fill
    if fill and fill.start_color and fill.start_color.rgb:
        rgb = str(fill.start_color.rgb)
        yellow_patterns = ["FFFF00", "FFFFE0", "FFF000", "FFCC00", "FFD700", "FFEB9C"]
        return any(pattern in rgb for pattern in yellow_patterns)
    return False


def hash_url(url):
    if not url:
        return None
    return hashlib.sha256(str(url).encode()).hexdigest()[:16]


def parse_xlsx(filepath, batch_id):
    """Parse XLSX and import all data. Returns (total, imported, flagged)."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    total = ws.max_row - 1  # minus header
    imported = 0
    flagged = 0

    # Build header map to locate columns
    header_map = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col_idx).value
        if val:
            header_map[str(val).strip().lower()] = col_idx

    for row_idx in range(2, ws.max_row + 1):
        try:
            row_data = {}
            for col_idx in range(1, ws.max_column + 1):
                row_data[col_idx] = ws.cell(row=row_idx, column=col_idx).value

            def safe_str(val):
                return str(val).strip() if val is not None else ""

            nome = safe_str(row_data.get(header_map.get("nome completo", 5)))
            escola_nome = safe_str(row_data.get(header_map.get("escola", 9)))
            cidade = safe_str(row_data.get(header_map.get("município", 8)))
            estado = safe_str(row_data.get(header_map.get("estado", 7)))
            polo = safe_str(row_data.get(header_map.get("polo", 10)))
            xlsx_id = safe_str(row_data.get(header_map.get("id", 1)))
            acertos_val = row_data.get(header_map.get("acertos", 6))
            if acertos_val is not None:
                try:
                    acertos = int(acertos_val)
                except (ValueError, TypeError):
                    acertos = None
            else:
                acertos = None

            justificativa_val = safe_str(row_data.get(header_map.get("justificativa", 4)))
            avaliador = safe_str(row_data.get(header_map.get("avaliador", 2)))
            status_xlsx = safe_str(row_data.get(header_map.get("status", 3)))
            situacao = safe_str(row_data.get(header_map.get("situação", 11)))

            if not nome:
                continue

            # Get or create school
            school = School.query.filter_by(name=escola_nome).first()
            if not school:
                school = School(name=escola_nome, city=cidade, state=estado, polo=polo)
                db.session.add(school)
                db.session.flush()

            # Don't auto-reject - fraud detection algorithm handles this
            status_student = "pending"

            student = Student(
                xlsx_id=xlsx_id,
                name=nome,
                school_id=school.id,
                score=acertos,
                status=status_student,
                xlsx_status=status_xlsx,
                xlsx_avaliador=avaliador,
                xlsx_situacao=situacao,
                raw_xlsx_row=row_idx,
                import_batch_id=batch_id,
            )
            db.session.add(student)
            db.session.flush()

            # Import answers for Q1..Q20
            for q_num in range(1, 21):
                col_name = f"q{q_num}"
                col_idx = header_map.get(col_name)
                if col_idx:
                    option_raw = row_data.get(col_idx)
                    option = str(option_raw).strip().upper() if option_raw is not None else ""
                    option = option[:1] if option else None
                    if option and option in "ABCDE":
                        answer = Answer(
                            student_id=student.id,
                            question_number=q_num,
                            selected_option=option,
                        )
                        db.session.add(answer)

            # Import justification
            if justificativa_val and justificativa_val.lower() not in ("não enviou", "nao enviou", "sem justificativa", "none"):
                justification = Justification(
                    student_id=student.id,
                    photo_url=justificativa_val,
                    photo_hash=hash_url(justificativa_val),
                )
                db.session.add(justification)

            # Check for yellow highlighting (fraud risk)
            is_yellow = False
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if is_yellow_fill(cell):
                    is_yellow = True
                    break

            if is_yellow:
                student.is_flagged = True
                student.flag_level = "high"
                fraud_flag = FraudFlag(
                    student_id=student.id,
                    source="manual",
                    level="high",
                    reason="Linha marcada em amarelo no XLSX — alto risco de fraude",
                )
                db.session.add(fraud_flag)
                flagged += 1

            imported += 1

        except Exception as e:
            print(f"Error importing row {row_idx}: {e}")
            db.session.rollback()
            continue

    db.session.commit()

    # Update import log
    batch = db.session.get(ImportLog, batch_id)
    if batch:
        batch.rows_total = total
        batch.rows_imported = imported
        batch.rows_flagged = flagged
        batch.status = "completed"
        db.session.commit()

    return total, imported, flagged
